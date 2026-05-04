from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

FILE_NAME = "finanse.xlsx"

def stworz_arkusz() -> tuple[Workbook, Worksheet]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Finanse"
    return wb, ws

def ustaw_wydatki(ws: Worksheet) -> None:
    ws["A1"] = "Czynsz"
    ws["B1"] = 100_000

    ws["A2"] = "Jedzenie"
    ws["B2"] = 650
    
def ustaw_sume(ws: Worksheet) -> None:
    ws["A3"] = "Suma"
    ws["B3"] = "=SUM(B1:B2)"
    
    
def main() -> None:
    wb, ws = stworz_arkusz()
    
    ustaw_wydatki(ws)
    ustaw_sume(ws)
    
    wb.save(FILE_NAME)


if __name__ == "__main__":
    main()

# WERSJA BARDZIEJ DYNAMICZNA
# from openpyxl import Workbook
# from openpyxl.worksheet.worksheet import Worksheet

# FILE_NAME = "finanse.xlsx"

# def stworz_arkusz() -> tuple[Workbook, Worksheet]:
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Finanse"
#     return wb, ws

# def dodaj_wydatki(ws: Worksheet, wydatki: list[tuple[str, int | float]]) -> None:
#     for wiersz in wydatki:
#         ws.append(wiersz)

# def dodaj_sume(ws: Worksheet, liczba_wydatkow: int) -> None:
#     ws[f"A{liczba_wydatkow + 1}"] = "Suma"
#     ws[f"B{liczba_wydatkow + 1}"] = f"=SUM(B1:B{liczba_wydatkow})"


# def main() -> None:
#     wb, ws = stworz_arkusz()
    
#     wydatki = [
#         ("Czynsz", 100_000),
#         ("Jedzenie", 650)
#     ]

#     dodaj_wydatki(ws, wydatki)
#     dodaj_sume(ws, len(wydatki))
    
#     wb.save(FILE_NAME)

  
# if __name__ == "__main__":
#     main()